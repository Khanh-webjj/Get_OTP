import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from datetime import datetime
print('Excel Handler');

def ReadExcel(path):
    print('Reading Excel From Path '+path)
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    row0 = list(sheet.rows)
    n = len(row0)
    details = []
    for row in list(sheet.rows)[1:]:
        j = 0
        map={}
        for cell in row:
            if j >= n:
                break;
            elif row0[j].value != None and cell.value!=None:
                map[row0[j].value] = cell.value
            j+=1;
        if map:
            details.append(map)
    return details

def Write_excel(path, contents):
    columns = list(contents[0].key())
    wb = openpyxl.Workbook()
    sheet = wb.active

    for i in range(1, len(columns)+1):
        sheet.cell(row=1, column=i).value = columns[i-1]
    i=2;
    for entry in contents:
        for j in range(1, len(columns)+1):
            sheet.cell(row=i, column=j).value = entry[columns[j-1]]
        i += 1;
    wb.save(path)

# read from excel
# def read

def append_to_excel(phone_number, status1, status2, message,idx,workbook):
    """
    Thêm một dòng mới vào file Excel với thông tin được cung cấp.

    Args:
        phone_number (str): Số điện thoại.
        status1 (str): Trạng thái 1.
        status2 (str): Trạng thái 2.
        timestamp (datetime): Thời gian để ghi vào file.
        file_name (str): Tên file Excel. Mặc định là 'data.xlsx'.

    """

    try:
        # Tải workbook và chọn sheet đầu tiên

        sheet = workbook.active

        # Lấy số hàng hiện tại để xác định vị trí dòng tiếp theo
        next_row = idx+1
        print("next_row",next_row)
        # Thêm dữ liệu vào dòng tiếp theo
        sheet.cell(row=next_row, column=1, value=phone_number)
        sheet.cell(row=next_row, column=2, value=status1)
        sheet.cell(row=next_row, column=3, value=status2)
        sheet.cell(row=next_row, column=4, value=message)
        sheet.cell(row=next_row, column=5, value= datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

        # Tự động điều chỉnh kích thước cột
        for col in range(1, 6):
            max_length = 0
            column = get_column_letter(col)
            for cell in sheet[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

        # Lưu lại workbook
        # workbook.save()
        print(f"Data appended successfully to .")

    except FileNotFoundError:
        print(f"File not found. Please check the file path.")
    except Exception as e:
        print(f"An error occurred: {e}")

def merge_excel_files(base_file, new_file):
    # Mở tệp gốc
    base_wb = openpyxl.load_workbook(base_file)
    base_ws = base_wb.active

    # for new_file in new_files:
        # Mở tệp mới
    new_wb = openpyxl.load_workbook(new_file)
    new_ws = new_wb.active

        # Sao chép nội dung từ tệp mới vào tệp gốc
    for row in new_ws.iter_rows(min_row=2, values_only=True):
        base_ws.append(row)

    # Lưu tệp gốc sau khi gộp
    base_wb.save(base_file)